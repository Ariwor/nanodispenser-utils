#!/usr/bin/env python3
"""
I.DOT Pipetting Scheme Generator

Reads an Excel template describing source plate layout and combinatorial
(or manual_rows/manual_columns) reaction design, then outputs a CSV ready for the I.DOT
nanodispenser software.

Usage:
    python generate_idot.py <input.xlsx> [output.csv]
"""

import csv
import io
import os
import sys
from datetime import datetime
from itertools import product as cartesian_product

import openpyxl


# ── Well helpers ────────────────────────────────────────────────────

ROWS_96 = "ABCDEFGH"
COLS_96 = range(1, 13)


def well_sort_key(well: str) -> tuple:
    row = well[0].upper()
    col = int(well[1:])
    return (col, ROWS_96.index(row))


def generate_target_wells(count: int) -> list[str]:
    """Return *count* wells in column-major order (A1, B1, ..., H1, A2, ...)."""
    wells = [f"{r}{c}" for c in COLS_96 for r in ROWS_96]
    if count > len(wells):
        raise ValueError(
            f"Need {count} target wells but a 96-well plate only has {len(wells)}."
        )
    return wells[:count]


# ── Cell parsing helper ──────────────────────────────────────────────

def _cell_to_str(val) -> str:
    """Safely convert any Excel cell value to a stripped string.

    Handles numeric reagent names like -1 or 0 that openpyxl returns
    as int/float rather than str.
    """
    if val is None:
        return ""
    return str(val).strip()


# ── Excel readers ───────────────────────────────────────────────────

def read_settings(ws) -> dict:
    settings = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        key, val = row
        if key is None:
            continue
        settings[_cell_to_str(key)] = val
    return settings


def read_source_plate(ws) -> dict[str, list[str]]:
    reagent_to_well: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        well, reagent = row
        w = _cell_to_str(well).upper()
        r = _cell_to_str(reagent)
        if not w or not r:
            continue
        reagent_to_well.setdefault(r, []).append(w)
    return reagent_to_well


def read_combinatorial(ws) -> tuple[list[str], list[list[str]]]:
    headers = [_cell_to_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    common_parts: list[str] = []
    groups: list[list[str]] = []

    for col_idx, header in enumerate(headers):
        if not header:
            continue
        col_values = []
        for row in ws.iter_rows(min_row=2, min_col=col_idx + 1,
                                max_col=col_idx + 1, values_only=True):
            s = _cell_to_str(row[0])
            if s:
                col_values.append(s)

        if header.lower() == "common":
            common_parts = col_values
        else:
            if col_values:
                groups.append(col_values)

    return common_parts, groups


def read_manual_rows(ws) -> list[tuple[str, list[str]]]:
    reactions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        target = _cell_to_str(vals[0]).upper()
        if not target:
            continue
        parts = [_cell_to_str(v) for v in vals[1:] if _cell_to_str(v)]
        if parts:
            reactions.append((target, parts))
    return reactions


def read_manual_columns(ws) -> list[tuple[str, list[str]]]:
    """Read a sheet where columns = target wells, rows = parts under each."""
    headers = [_cell_to_str(c.value).upper()
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
    reactions = []
    for col_idx, well in enumerate(headers):
        if not well:
            continue
        parts = []
        for row in ws.iter_rows(min_row=2, min_col=col_idx + 1,
                                max_col=col_idx + 1, values_only=True):
            s = _cell_to_str(row[0])
            if s:
                parts.append(s)
        if parts:
            reactions.append((well, parts))
    return reactions


# ── Reaction builders ───────────────────────────────────────────────

def build_combinatorial_reactions(
    common_parts: list[str],
    groups: list[list[str]],
) -> list[tuple[str, list[str]]]:
    """
    Return [(target_well, [part_names]), ...] from the Cartesian product
    of *groups*, with *common_parts* appended to every reaction.
    """
    if not groups:
        raise ValueError("No part groups provided for combinatorial mode.")

    combos = list(cartesian_product(*groups))
    targets = generate_target_wells(len(combos))

    reactions = []
    for target, combo in zip(targets, combos):
        parts = list(combo) + common_parts
        reactions.append((target, parts))
    return reactions


def build_manual_reactions(
    manual_rows: list[tuple[str, list[str]]],
) -> list[tuple[str, list[str]]]:
    if not manual_rows:
        raise ValueError("No reactions found in manual specification.")
    return manual_rows


# ── Dispensing logic ────────────────────────────────────────────────

def build_dispense_rows(
    reactions: list[tuple[str, list[str]]],
    reagent_to_wells: dict[str, list[str]],
    total_vol: float = 10.0,
    dispense_vol: float = 0.5,
    mm_well: str = "A1",
    mm_name: str = "Mastermix",
) -> list[tuple[str, str, float, str]]:
    """
    Return list of (source_well, target_well, volume_uL, liquid_name).

    Automatically balances load across multiple source wells when the
    same reagent is listed in more than one well on the source plate.
    """
    all_parts: set[str] = set()
    for _, parts in reactions:
        all_parts.update(parts)

    missing = all_parts - set(reagent_to_wells.keys())
    if missing:
        raise ValueError(
            "These parts are listed in reactions but not on the Source Plate:\n  "
            + ", ".join(sorted(missing))
        )

    well_usage: dict[str, int] = {}

    def pick_source_well(reagent: str) -> str:
        wells = reagent_to_wells[reagent]
        best = min(wells, key=lambda w: well_usage.get(w, 0))
        well_usage[best] = well_usage.get(best, 0) + 1
        return best

    rows: list[tuple[str, str, float, str]] = []
    for target, parts in reactions:
        n_parts = len(parts)
        mm_vol = round(total_vol - n_parts * dispense_vol, 4)
        if mm_vol < 0:
            raise ValueError(
                f"Reaction in {target} has {n_parts} parts x {dispense_vol} uL "
                f"= {n_parts * dispense_vol} uL, exceeding total volume {total_vol} uL."
            )

        for part in parts:
            src = pick_source_well(part)
            rows.append((src, target, dispense_vol, part))

        if mm_vol > 0:
            rows.append((mm_well, target, mm_vol, mm_name))

    return rows


# ── Nanodispenser input CSV writer ──────────────────────────────────────────────────────

def write_idot_csv(
    dispense_rows: list[tuple[str, str, float, str]],
    output_path: str,
    experiment_name: str = "Nanodispenser_Demo",
    user_name: str = "",
    source_plate_type: str = "S.100 Plate",
    target_plate_type: str = "MWP 96",
):
    now = datetime.now()
    header_lines = [
        [experiment_name, "1.9.0.3", user_name,
         now.strftime("%d.%m.%Y"), now.strftime("%H:%M:%S"), "", "", ""],
        [source_plate_type, "Source Plate 1", "", "0.00008",
         target_plate_type, "Target Plate 1", "", "Waste Tube"],
        [
            "DispenseToWaste=True",
            "DispenseToWasteCycles=3",
            "DispenseToWasteVolume=1e-7",
            "UseDeionisation=True",
            "OptimizationLevel=ReorderAndParallel",
            "WasteErrorHandlingLevel=Ask",
            "SaveLiquids=Never",
            "",
        ],
        ["Source Well", "Target Well", "Volume [uL]", "Liquid Name", "", "", "", ""],
    ]

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        for line in header_lines:
            writer.writerow(line)
        for src, tgt, vol, name in dispense_rows:
            writer.writerow([src, tgt, vol, name, "", "", "", ""])


# ── Summary ─────────────────────────────────────────────────────────

def print_summary(reactions, dispense_rows, total_vol, dispense_vol,
                   summary_path: str | None = None):
    vol_per_source: dict[str, float] = {}
    name_for_source: dict[str, str] = {}
    for src, _, vol, name in dispense_rows:
        vol_per_source[src] = vol_per_source.get(src, 0) + vol
        name_for_source[src] = name

    n_parts = len(reactions[0][1]) if reactions else 0
    mm_vol = round(total_vol - n_parts * dispense_vol, 4) if reactions else 0
    target_wells = sorted([t for t, _ in reactions], key=well_sort_key)

    print(f"\n{'=' * 50}")
    print(f"  SUMMARY")
    print(f"{'=' * 50}")
    print(f"  Total reactions:       {len(reactions)}")
    print(f"  Parts per reaction:    {n_parts}")
    print(f"  Mastermix per rxn:     {mm_vol} uL")
    print(f"  Target wells:          {target_wells[0]} .. {target_wells[-1]}")

    print(f"\n  {'Source Well':<14}{'Reagent':<18}{'Total Vol (uL)':>14}")
    print(f"  {'-' * 46}")
    for well in sorted(vol_per_source, key=well_sort_key):
        name = name_for_source[well]
        vol = round(vol_per_source[well], 2)
        print(f"  {well:<14}{name:<18}{vol:>14.1f}")

    print(f"\n  NOTE: Add dead volume (~3 uL per well for S.100)")
    print(f"        on top of these values.")
    print(f"{'=' * 50}\n")

    if summary_path:
        with open(summary_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Total reactions", len(reactions)])
            writer.writerow(["Parts per reaction", n_parts])
            writer.writerow(["Mastermix per rxn (uL)", mm_vol])
            writer.writerow(["Target wells", f"{target_wells[0]} .. {target_wells[-1]}"])
            writer.writerow([])
            writer.writerow(["Source Well", "Reagent", "Total Vol (uL)"])
            for well in sorted(vol_per_source, key=well_sort_key):
                name = name_for_source[well]
                vol = round(vol_per_source[well], 2)
                writer.writerow([well, name, vol])
        print(f"  Summary written to: {summary_path}")


# ── Main ────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_idot.py <input.xlsx> [output.csv]")
        print("\nExample:")
        print("  python generate_idot.py template.xlsx")
        print("  python generate_idot.py template.xlsx my_output.csv")
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.isfile(input_path):
        sys.exit(f"Error: file not found: {input_path}")

    output_path = (
        sys.argv[2] if len(sys.argv) >= 3
        else f"{os.path.splitext(os.path.basename(input_path))[0]}_idot.csv"
    )

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    for s in ("Settings", "Source Plate"):
        if s not in wb.sheetnames:
            sys.exit(f"Error: missing required sheet '{s}' in {input_path}")

    settings = read_settings(wb["Settings"])
    reagent_to_wells = read_source_plate(wb["Source Plate"])

    mode = str(settings.get("Mode", "combinatorial")).strip().lower()
    total_vol = float(settings.get("Total Reaction Volume (uL)", 10))
    dispense_vol = float(settings.get("Part Dispense Volume (uL)", 0.5))
    mm_well = str(settings.get("Mastermix Source Well", "A1")).strip().upper()
    mm_name = str(settings.get("Mastermix Name", "Mastermix")).strip()

    try:
        if mode == "combinatorial":
            if "Combinatorial" not in wb.sheetnames:
                sys.exit("Error: mode is 'combinatorial' but no 'Combinatorial' sheet found.")
            common_parts, groups = read_combinatorial(wb["Combinatorial"])
            reactions = build_combinatorial_reactions(common_parts, groups)
        elif mode == "manual_columns":
            if "Manual Columns" not in wb.sheetnames:
                sys.exit("Error: mode is 'manual_columns' but no 'Manual Columns' sheet found.")
            reactions = build_manual_reactions(read_manual_columns(wb["Manual Columns"]))
        elif mode == "manual_rows":
            if "Manual" not in wb.sheetnames:
                sys.exit("Error: mode is 'manual_rows' but no 'Manual Rows' sheet found.")
            reactions = build_manual_reactions(read_manual_rows(wb["Manual Rows"]))
        else:
            sys.exit(f"Error: unknown mode '{mode}'. Use 'combinatorial', 'manual_rows', or 'manual_columns'.")

        dispense_rows = build_dispense_rows(
            reactions, reagent_to_wells,
            total_vol=total_vol, dispense_vol=dispense_vol,
            mm_well=mm_well, mm_name=mm_name,
        )
    except ValueError as e:
        sys.exit(f"Error: {e}")

    write_idot_csv(
        dispense_rows, output_path,
        experiment_name=str(settings.get("Experiment Name", "Nanodispenser_Demo")),
        user_name=str(settings.get("User Name", "") or ""),
        source_plate_type=str(settings.get("Source Plate Type", "S.100 Plate")),
        target_plate_type=str(settings.get("Target Plate Type", "MWP 96")),
    )

    summary_path = output_path.replace("_idot.csv", "_summary.csv")
    if summary_path == output_path:
        summary_path = os.path.splitext(output_path)[0] + "_summary.csv"

    print(f"\n  Nanodispenser input CSV written to: {output_path}")
    print_summary(reactions, dispense_rows, total_vol, dispense_vol,
                  summary_path=summary_path)


if __name__ == "__main__":
    main()
